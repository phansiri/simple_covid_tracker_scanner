from django.shortcuts import render, redirect
from django.http import HttpResponseRedirect, HttpResponse
from django.urls import reverse
from django.utils import timezone
from datetime import datetime
from .models import Person, Event, Location
from cac_parser import PDF417Barcode
from .forms import EventForm, PersonForm, LocationForm
from openpyxl import Workbook


def event_list(request):
    events = Event.objects.all().order_by('-arrived')
    loc = Location.objects.exists()
    print(loc)

    if loc is False:
        return render(request, 'covid_app/set_location.html')
    else:
        loc = Location.objects.all()

    todays_date = datetime.now()

    context = {
        'events': events,
        'loc': loc,
        'date': todays_date
    }
    return render(request, 'covid_app/event_list.html', context)


def reports(request):
    return render(request, 'covid_app/reports.html', {})


def download_report(request):
    events = Event.objects.all().order_by('-arrived')

    timeframe = datetime.now()

    dls = [event for event in events if event.arrived.date() == timeframe.date()]

    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = f'attachment; filename=Covid-Tracker-{timeframe.date()}-{timeframe.hour}:{timeframe.minute}.xlsx'
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Covid_Tracker'
    headers = ['edipi', 'rank', 'lastname', 'firstname', 'branch', 'category', 'location', 'date', 'time']
    row_num = 1

    for col_num, column_title in enumerate(headers, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    for dl in dls:
        row_num += 1

        row = [
            dl.initiator.edipi,
            dl.initiator.rank,
            dl.initiator.lname,
            dl.initiator.fname,
            dl.initiator.branch,
            dl.initiator.category,
            dl.location.location,
            dl.arrived.date(),
            dl.arrived.time()
        ]

        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    column_dimensions = worksheet.column_dimensions['H']
    column_dimensions.width = 10

    workbook.save(response)
    return response


def set_location(request):
    if request.method.lower() == 'post':
        location = request.POST.get('location')
        form = LocationForm(request.POST)
        if form.is_valid():
            loc = form.save(commit=False)
            loc.location = location
            loc.save()
            return redirect('event_list')


def scan(request):
    if request.method.lower() == 'post':
        barcode = request.POST.get('scanner')
        loc_pk = request.POST.get('location')

        item = PDF417Barcode(barcode)  # parse cac for vital information
        person: Person = None
        print(item)
        try:
            person = Person.objects.get(edipi=item.edipi)
        except:
            pass

        if not person:  # person doesn't exist
            person = Person.objects.create(
                rank=item.rank.strip(),
                lname=item.lname.replace(".", "").strip(),
                fname=item.fname.replace(".", "").strip(),
                branch=item.branch.strip(),
                category=item.category.strip(),
                edipi=item.edipi,
            )
        location = Location.objects.get(pk=loc_pk)

        Event.objects.create(initiator=person, location=location)

        return redirect('event_list')
